"""
reporting/report_generator.py — PDF, Text & Excel Report-Generierung.

Erstellt professionelle Reports aus den Analyse-Ergebnissen:
- Text-Report: Maschinell lesbar, Git-versionierbar, schnell
- Excel-Report: Business-User friendly, interaktiv filterbar

Design: Report-Layer kennt nur KPIs + DataFrames, keine Business-Logik.
Saubere Trennung Analyse ↔ Präsentation (MVC-ähnlich).
"""

import logging
from datetime import datetime
from pathlib import Path
from typing import Optional

import pandas as pd

from config import ExportConfig, PipelineConfig

logger = logging.getLogger(__name__)


# ─────────────────────────────────────────────
# TEXT REPORT
# ─────────────────────────────────────────────

def generate_text_report(
    kpis: dict,
    kpi_formatted: dict,
    insights: str,
    monthly_df: pd.DataFrame,
    products_df: pd.DataFrame,
    rfm_df: pd.DataFrame,
    config: PipelineConfig,
    output_dir: Path
) -> Path:
    """
    Generiert strukturierten Text-Report im professionellen Format.

    Layout analog zu Unternehmens-Management-Reports:
    Header → Executive Summary → KPIs → Insights → Empfehlungen → Daten-Anhang

    Args:
        kpis: KPI-Dictionary (Rohdaten)
        kpi_formatted: Formatierte KPI-Strings (für Display)
        insights: AI-generierter oder Fallback-Insight-Text
        monthly_df: Monatliche Umsätze
        products_df: Produkt-Performance
        rfm_df: RFM-Segmentierung
        config: Pipeline-Konfiguration
        output_dir: Ausgabe-Pfad

    Returns:
        Pfad zur generierten Report-Datei
    """
    now = datetime.now()
    date_str = now.strftime("%Y-%m-%d")
    time_str = now.strftime("%H:%M:%S")

    period = kpis.get("period", {})
    rev = kpis.get("revenue", {})
    growth = kpis.get("growth", {})
    customers = kpis.get("customers", {})
    top_perf = kpis.get("top_performers", {})
    products_meta = kpis.get("products", {})

    # ── Monatliche Top/Flop Zeilen ─────────────────────────────
    monthly_sorted = monthly_df.copy()
    monthly_sorted["month"] = pd.to_datetime(monthly_sorted["month"])
    monthly_sorted = monthly_sorted.sort_values("revenue", ascending=False)

    top3_months = monthly_sorted.head(3)
    flop3_months = monthly_sorted.tail(3)

    # ── Produkt-Tabelle (Top 10) ───────────────────────────────
    product_rows = ""
    for _, row in products_df.head(10).iterrows():
        abc = row.get("abc_class", "-")
        product_rows += (
            f"  {row['name']:<28} "
            f"{row['category']:<14} "
            f"€{row['total_revenue']:>10,.0f}  "
            f"{row['margin_pct']:>5.1f}%  "
            f"{abc}\n"
        )

    # ── RFM-Zusammenfassung ────────────────────────────────────
    rfm_summary = ""
    if "rfm_segment" in rfm_df.columns:
        for seg, count in rfm_df["rfm_segment"].value_counts().items():
            pct = count / len(rfm_df) * 100
            rfm_summary += f"  {seg:<22}: {count:>4} Kunden ({pct:>4.1f}%)\n"

    # ── Report Text zusammenbauen ──────────────────────────────
    border = "═" * 65
    thin_line = "─" * 65

    yoy_str = f"{growth.get('yoy_pct', 0):+.1f}%" if growth.get("yoy_pct") is not None else "N/A"
    mom_str = f"{growth.get('avg_mom_pct_last6m', 0):+.1f}%" if growth.get("avg_mom_pct_last6m") is not None else "N/A"
    margin_str = f"{rev.get('gross_margin_pct', 0):.1f}%" if rev.get("gross_margin_pct") is not None else "N/A"

    report_text = f"""{border}
{config.company_name.upper()} — {config.report_title.upper()}
Zeitraum: {period.get('start', 'N/A')} bis {period.get('end', 'N/A')}
Generiert: {date_str} {time_str}
{border}

EXECUTIVE SUMMARY
{thin_line}
Gesamtumsatz (Total):    €{rev.get('total', 0):>12,.2f}
Umsatz (letzte 12M):     €{rev.get('last_12_months', 0):>12,.2f}  ({yoy_str} YoY)
Bruttomarge:              {margin_str:>12}
Ø MoM-Wachstum (6M):     {mom_str:>12}
Gesamtbestellungen:       {rev.get('total_orders', 0):>12,}
Aktive Kunden:            {customers.get('active_total', 0):>12,}
Ø Bestellwert:           €{rev.get('avg_order_value', 0):>12,.2f}

REGIONALE PERFORMANCE
{thin_line}
Beste Region:    {top_perf.get('best_region', 'N/A'):<20} €{top_perf.get('best_region_revenue', 0):>12,.0f}
Schwächste Region: {top_perf.get('worst_region', 'N/A'):<18} €{top_perf.get('worst_region_revenue', 0):>12,.0f}
Bester Kanal:    {top_perf.get('best_channel', 'N/A')}

PRODUKT-PERFORMANCE (Top 10 nach Umsatz)
{thin_line}
  {'Produkt':<28} {'Kategorie':<14} {'Umsatz':>12}  {'Marge':>5}  ABC
  {'-'*28} {'-'*14} {'-'*12}  {'-'*5}  ---
{product_rows}
  Legende: A=Top 80% Umsatz | B=Nächste 15% | C=Letzte 5%

KUNDEN-SEGMENTIERUNG (RFM-Analyse)
{thin_line}
  Champions-Anteil: {customers.get('champions_pct', 0):.1f}% aller Kunden
  At-Risk-Kunden:   {customers.get('at_risk_count', 0):,} (Churn-Risiko)

{rfm_summary}
UMSATZ-HIGHLIGHTS
{thin_line}
Top-3-Monate:
"""

    for _, row in top3_months.iterrows():
        report_text += f"  {row['month'].strftime('%B %Y'):<18} €{row['revenue']:>12,.0f}  ({row.get('mom_growth_pct', 0):+.1f}% MoM)\n"

    report_text += f"""
Flop-3-Monate:
"""
    for _, row in flop3_months.iterrows():
        report_text += f"  {row['month'].strftime('%B %Y'):<18} €{row['revenue']:>12,.0f}  ({row.get('mom_growth_pct', 0):+.1f}% MoM)\n"

    report_text += f"""
{border}
KI-GENERIERTE INSIGHTS & EMPFEHLUNGEN
{thin_line}

{insights}

{border}
TECHNISCHE DETAILS
{thin_line}
Datenquelle:    SQLite Datenbank (sales.db)
Analysemethode: Python / Pandas / RFM-Modell
Visualisierung: Matplotlib / Seaborn (5 Charts)
KI-Modell:      OpenAI GPT / Fallback-Modus
Pipeline-Stack: Python 3.10+ | Pandas | NumPy
{border}
Ende des Reports — {config.company_name} | {date_str}
{border}
"""

    # ── Speichern ──────────────────────────────────────────────
    output_dir.mkdir(parents=True, exist_ok=True)
    filename = config.export.text_filename.format(date=date_str)
    output_path = output_dir / filename

    with open(output_path, "w", encoding="utf-8") as f:
        f.write(report_text)

    logger.info(f"Text-Report gespeichert: {output_path.name} ({output_path.stat().st_size / 1024:.1f} KB)")
    return output_path


# ─────────────────────────────────────────────
# EXCEL REPORT
# ─────────────────────────────────────────────

def generate_excel_report(
    kpis: dict,
    monthly_df: pd.DataFrame,
    products_df: pd.DataFrame,
    region_channel_df: pd.DataFrame,
    rfm_df: pd.DataFrame,
    customer_stats_df: pd.DataFrame,
    config: PipelineConfig,
    output_dir: Path
) -> Path:
    """
    Generiert mehrseitigen Excel-Report mit Formatierung.

    Sheets:
    1. KPI Overview — Kennzahlen-Übersicht mit bedingter Formatierung
    2. Monatliche Umsätze — Vollständige Zeitreihe
    3. Produkt-Ranking — ABC-klassifiziert mit Marge
    4. Regionale Matrix — Pivot Kanal × Region
    5. Kundensegmente — RFM-Zusammenfassung

    Args:
        kpis: KPI-Dictionary
        monthly_df: Monatliche Umsätze
        products_df: Produkt-Performance
        region_channel_df: Region × Kanal
        rfm_df: RFM-Segmentierung
        customer_stats_df: Kundenstamm-Daten
        config: Pipeline-Konfiguration
        output_dir: Ausgabe-Pfad

    Returns:
        Pfad zur generierten Excel-Datei
    """
    try:
        import openpyxl
        from openpyxl.styles import (
            Font, PatternFill, Alignment, Border, Side, numbers
        )
        from openpyxl.utils import get_column_letter
        from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
    except ImportError:
        logger.warning("openpyxl nicht installiert — Excel-Export übersprungen")
        return None

    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / config.export.excel_filename

    # Farben (Corporate)
    HEADER_BG = "1B4F72"
    HEADER_FG = "FFFFFF"
    ALT_ROW_BG = "EBF5FB"
    ACCENT_BG = "2E86C1"
    GREEN = "27AE60"
    RED   = "922B21"

    def header_style(ws, row: int, values: list, bg_color: str = HEADER_BG):
        """Schreibt eine Header-Zeile mit Formatierung."""
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = Font(bold=True, color=HEADER_FG, name="Calibri", size=10)
            cell.fill = PatternFill(fill_type="solid", fgColor=bg_color)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def auto_width(ws, min_width: int = 12, max_width: int = 40):
        """Setzt optimale Spaltenbreite."""
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=0)
            ws.column_dimensions[get_column_letter(col[0].column)].width = (
                max(min_width, min(max_len + 3, max_width))
            )

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # ── Sheet 1: KPI Overview ──────────────────────────────
        rev = kpis.get("revenue", {})
        growth = kpis.get("growth", {})
        customers = kpis.get("customers", {})
        top_perf = kpis.get("top_performers", {})
        period = kpis.get("period", {})

        kpi_data = {
            "KPI": [
                "Gesamtumsatz", "Umsatz (letzte 12M)", "Umsatz (Vorjahr 12M)",
                "YoY-Wachstum", "Ø MoM-Wachstum (6M)", "Bruttomarge",
                "Gesamtbestellungen", "Aktive Kunden", "Ø Bestellwert",
                "Champions-Kunden", "At-Risk-Kunden",
                "Beste Region", "Schwächste Region", "Bester Kanal",
                "Top-Produkt",
            ],
            "Wert": [
                rev.get("total", 0), rev.get("last_12_months", 0), rev.get("prev_12_months", 0),
                growth.get("yoy_pct"), growth.get("avg_mom_pct_last6m"),
                rev.get("gross_margin_pct"),
                rev.get("total_orders", 0), customers.get("active_total", 0),
                rev.get("avg_order_value", 0),
                customers.get("champions_count", 0), customers.get("at_risk_count", 0),
                top_perf.get("best_region"), top_perf.get("worst_region"),
                top_perf.get("best_channel"),
                kpis.get("products", {}).get("top_product_name"),
            ],
            "Einheit": [
                "€", "€", "€", "%", "%", "%",
                "Stück", "Kunden", "€",
                "Kunden", "Kunden",
                "", "", "",
                "",
            ],
            "Zeitraum": [
                f"{period.get('start')} – {period.get('end')}",
                "Letzte 12 Monate", "Vorjahr 12 Monate",
                "YoY", "Letzte 6 Monate", "Gewichtet",
                "Gesamt", "Gesamt", "Durchschnitt",
                "Aktuell", "Aktuell",
                "", "", "", "",
            ]
        }

        kpi_df = pd.DataFrame(kpi_data)
        kpi_df.to_excel(writer, sheet_name="KPI Overview", index=False)

        # Formatierung
        ws = writer.sheets["KPI Overview"]
        header_style(ws, 1, ["KPI", "Wert", "Einheit", "Zeitraum"])
        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 22

        # Bedingte Formatierung für Wachstumsraten (Zeile 4+5)
        for row_idx in [4, 5]:
            cell = ws.cell(row=row_idx + 1, column=2)
            if isinstance(cell.value, (int, float)) and cell.value is not None:
                color = GREEN if cell.value >= 0 else RED
                cell.fill = PatternFill(fill_type="solid", fgColor=color + "20")
                cell.font = Font(bold=True, color=color)

        auto_width(ws)

        # ── Sheet 2: Monatliche Umsätze ────────────────────────
        monthly_export = monthly_df.copy()
        monthly_export["month"] = pd.to_datetime(monthly_export["month"]).dt.strftime("%Y-%m")

        monthly_export.to_excel(writer, sheet_name="Monatliche Umsätze", index=False)
        ws2 = writer.sheets["Monatliche Umsätze"]
        header_style(ws2, 1, list(monthly_export.columns), bg_color=ACCENT_BG)
        ws2.freeze_panes = "A2"
        auto_width(ws2)

        # ── Sheet 3: Produkt-Ranking ───────────────────────────
        product_export = products_df.copy()
        export_cols = [c for c in [
            "name", "category", "price", "margin_pct",
            "total_revenue", "total_units", "total_orders",
            "abc_class"
        ] if c in product_export.columns]
        product_export = product_export[export_cols].copy()
        product_export.columns = [
            c.replace("_", " ").title() for c in product_export.columns
        ]

        product_export.to_excel(writer, sheet_name="Produkt-Ranking", index=False)
        ws3 = writer.sheets["Produkt-Ranking"]
        header_style(ws3, 1, list(product_export.columns))
        ws3.freeze_panes = "A2"
        auto_width(ws3)

        # Bedingte Formatierung: Umsatz-Spalte (Datenbalken)
        rev_col = next(
            (i for i, c in enumerate(product_export.columns, 1) if "Revenue" in c),
            None
        )
        if rev_col:
            rev_col_letter = get_column_letter(rev_col)
            ws3.conditional_formatting.add(
                f"{rev_col_letter}2:{rev_col_letter}{len(product_export)+1}",
                DataBarRule(
                    start_type="min", end_type="max",
                    color="2E86C1"
                )
            )

        # ── Sheet 4: Regionale Matrix ──────────────────────────
        pivot = region_channel_df.pivot_table(
            index="region", columns="channel",
            values="revenue", aggfunc="sum", fill_value=0
        ).round(2)
        pivot["Gesamt"] = pivot.sum(axis=1)
        pivot.loc["GESAMT"] = pivot.sum()
        pivot = pivot.reset_index()
        pivot.columns.name = None

        pivot.to_excel(writer, sheet_name="Regionale Matrix", index=False)
        ws4 = writer.sheets["Regionale Matrix"]
        header_style(ws4, 1, list(pivot.columns))
        ws4.freeze_panes = "B2"
        auto_width(ws4)

        # ── Sheet 5: Kundensegmente ────────────────────────────
        if "rfm_segment" in rfm_df.columns:
            rfm_summary = rfm_df.groupby("rfm_segment").agg(
                Kunden_Anzahl=("customer_id", "count"),
                Ø_Umsatz=("monetary", "mean"),
                Ø_Bestellungen=("frequency", "mean"),
                Ø_Recency_Tage=("recency", "mean"),
                Gesamt_Umsatz=("monetary", "sum")
            ).round(2).reset_index()
            rfm_summary.columns = [c.replace("_", " ") for c in rfm_summary.columns]
            rfm_summary.to_excel(writer, sheet_name="Kundensegmente", index=False)
            ws5 = writer.sheets["Kundensegmente"]
            header_style(ws5, 1, list(rfm_summary.columns))
            auto_width(ws5)

    logger.info(f"Excel-Report gespeichert: {output_path.name} ({output_path.stat().st_size / 1024:.1f} KB)")
    return output_path
